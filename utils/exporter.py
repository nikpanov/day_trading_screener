import os
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from utils.logger import setup_logger
from emailer.notify import send_email_with_attachment
from datetime import datetime
from pytz import timezone as pytz_timezone

ET = pytz_timezone("US/Eastern")
logger = setup_logger()

def now_et():
    return datetime.now(ET)

def export_screener_results_to_excel(rows, run_timestamp):
    os.makedirs("output/screener_results", exist_ok=True)
    df = pd.DataFrame(rows)

    expected_columns = [
        "symbol", "company_name", "price", "rsi14", "ema20", "ema50", "vwap",
        "is_bullish", "failure_reason", "timestamp", "bullish_duration"
    ]
    for col in expected_columns:
        if col not in df.columns:
            df[col] = None
    df = df[expected_columns]

    # Remove tzinfo
    if "timestamp" in df.columns:
        df["timestamp"] = pd.to_datetime(df["timestamp"]).dt.tz_localize(None)

    run_timestamp_naive = run_timestamp.replace(tzinfo=None) if run_timestamp.tzinfo else run_timestamp
    filename = f"output/screener_results/screener_results_{run_timestamp_naive.strftime('%Y-%m-%d_%H%M')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Screener Results"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
            continue
        is_bullish = df.loc[r_idx - 2, "is_bullish"]
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if is_bullish else PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for cell in ws[r_idx]:
            cell.fill = fill

    try:
        wb.save(filename)
        logger.info(f"Exported screener results to Excel: {filename}")
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        return

    # Email Summary
    try:
        total_scanned = len(df)
        bullish_df = df[df["is_bullish"] == True]
        bullish_count = df["is_bullish"].sum()

        bullish_text = (
            "\n🟢 Bullish tickers:\n" +
            "\n".join(f"  - {row['symbol']}: ${row['price']} ({row['company_name']})" for _, row in bullish_df.iterrows())
        ) if not bullish_df.empty else "\n🟡 No bullish tickers detected."

        long_term_bulls = sum(1 for r in rows if r.get("bullish_duration", 0) >= 2)
        body = (
            f"📊 Screener run completed.\n\n"
            f"✅ Bullish tickers found: {bullish_count}\n"
            f"🔍 Total tickers scanned: {total_scanned}\n"
            f"{bullish_text}\n"
            f"⏳ Tickers bullish ≥ 2 days: {long_term_bulls}"
        )

        send_email_with_attachment("📈 Screener Results Available", body, filename)
    except Exception as e:
        logger.error(f"Failed to send email notification: {e}")

def export_backtest_results_to_excel(results, run_timestamp):
    os.makedirs("output/backtest_results", exist_ok=True)

    df = pd.DataFrame(results, columns=[
        "symbol", "signal_date", "buy_price", "sell_price",
        "sell_date", "gain_pct", "holding_days", "result"
    ])

    run_timestamp_naive = run_timestamp.replace(tzinfo=None) if run_timestamp.tzinfo else run_timestamp
    filename = f"output/backtest_results/backtest_results_{run_timestamp_naive.strftime('%Y-%m-%d_%H%M')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Backtest Results"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)
        if r_idx == 1:
            for cell in ws[r_idx]:
                cell.font = Font(bold=True)
            continue
        result = df.loc[r_idx - 2, "result"]
        fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if result == 'win' else                PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") if result == 'loss' else None
        if fill:
            for cell in ws[r_idx]:
                cell.fill = fill

    try:
        wb.save(filename)
        logger.info(f"Exported backtest results to Excel: {filename}")
    except Exception as e:
        logger.error(f"Failed to save backtest Excel file: {e}")